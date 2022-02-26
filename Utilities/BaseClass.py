import inspect
import logging
import pdb
import time

import pyautogui
import pytest
import openpyxl
from selenium.common.exceptions import ElementNotVisibleException, ElementNotSelectableException
from selenium.webdriver import ActionChains, Keys, TouchActions
from selenium.webdriver.common.by import By
from selenium.webdriver.support import expected_conditions
from selenium.webdriver.support.select import Select
from selenium.webdriver.support.wait import WebDriverWait

from PageObjects.Practise_form import Practise_form_po


@pytest.mark.usefixtures("init_browser")
class Baseclass:

    def Loggingfunc(self):
        logname = inspect.stack()[1][3]
        logger = logging.getLogger(logname)
        filehndler = logging.FileHandler('logfile.log')
        formatter = logging.Formatter("%(asctime)s : %(levelname)s : %(name)s : %(message)s")
        filehndler.setFormatter(formatter)
        logger.addHandler(filehndler)
        logger.setLevel(logging.DEBUG)
        return logger

    def Scrolldown(self, y):
        self.driver.execute_script(f"window.scrollTo(0,{y})")

    def Explicit_wait_by_clickable(self, locator):
        wait = WebDriverWait(self.driver, 10 , poll_frequency=1, ignored_exceptions=[ElementNotVisibleException, ElementNotSelectableException])
        wait.until(expected_conditions.element_to_be_clickable(locator))

    def Excel_data_drive(self, catogary):
        wb = openpyxl.load_workbook(".\\Datas\\Data_set.xlsx")
        sheet1 = wb.active
        if catogary == "name":
            return str(sheet1.cell(row=1, column=2).value)
        elif catogary == "Email":
            return str(sheet1.cell(row=2, column=2).value)
        elif catogary == "Current address":
            return str(sheet1.cell(row=3, column=2).value)
        elif catogary == "Permanent address":
            return str(sheet1.cell(row=4, column=2).value)

    def window_change(self, window_no):
        wh = self.driver.window_handles
        self.driver.switch_to.window(wh[window_no])

    def Explicit_wait_by_visiblity(self, locator):
        wait = WebDriverWait(self.driver, 15)
        wait.until(expected_conditions.visibility_of_element_located(locator))

    def scroll_vertical(self, y_value):
        self.driver.execute_script(f'scrollTo(0,{y_value})')

    def dropdown_byvalue(self, value):
        actions = ActionChains(self.driver)
        actions.send_keys(value).perform()
        time.sleep(2)
        actions.send_keys(Keys.ENTER).perform()

    def Select_by_dropdown(self, locator, val):
        s = Select(self.driver.find_element(*locator))
        s.select_by_value(f"{str(val)}")

    def date_picker_by_value(self, day, month, year, month_locator, year_locator):
        self.Select_by_dropdown(month_locator, month - 1)
        self.Select_by_dropdown(year_locator, year)
        if len(str(day)) == 1:
            day = str(day).zfill(2)
        else:
            pass
        self.driver.find_element(By.CSS_SELECTOR, f".react-datepicker__day.react-datepicker__day--0{str(day)}").click()

    def wait_for_alert(self, time):
        wait = WebDriverWait(self.driver, time)
        wait.until(expected_conditions.alert_is_present())

    def frame_wait_switch(self, duration, locator):
        wait = WebDriverWait(self.driver, duration)
        wait.until(expected_conditions.frame_to_be_available_and_switch_to_it(locator))

    def Explicit_wait_by_presence(self, locator):
        wait = WebDriverWait(self.driver, 10)
        wait.until(expected_conditions.presence_of_element_located(locator))

    def Explicit_wait_by_text_visiblity(self, locator, text):
        wait = WebDriverWait(self.driver, 10)
        wait.until(expected_conditions.text_to_be_present_in_element(locator, text))

    def Check_doc_ready_state(self):
        while self.driver.execute_script('return document.readyState') == 'complete':
            time.sleep(1)

    def date_picker_by_datetime(self, day, month, year, user_time, ele_locator):
        Act = ActionChains(self.driver)
        self.driver.find_element(*ele_locator).click()
        pyautogui.hotkey("ctrl","a")
        pyautogui.press("backspace")
        Act.send_keys(month+" "+str(day)+","+str(year)+" "+user_time).perform()
        Act.send_keys(Keys.ENTER).perform()

    def Select_dropdown_visible_text(self, locator, val):
        s = Select(self.driver.find_element(*locator))
        s.select_by_visible_text(f"{val}")

    def drag_vertical_with_yaxis(self, yaxis, locator):
        Ele = self.driver.find_element(*locator)
        ActionChains(self.driver).click_and_hold(Ele).drag_and_drop_by_offset(Ele, 0, yaxis).perform()

    def drag_vertical_with_xaxis(self, xaxis, locator):
        Ele = self.driver.find_element(*locator)
        ActionChains(self.driver).click_and_hold(Ele).drag_and_drop_by_offset(Ele, xaxis, 0).perform()

    def drag_both_axis(self, xaxis, yaxis, locator):
        Ele = self.driver.find_element(*locator)
        ActionChains(self.driver).click_and_hold(Ele).drag_and_drop_by_offset(Ele, xaxis, yaxis).perform()


